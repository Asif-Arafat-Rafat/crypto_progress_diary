from datetime import date
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
def file_name(coins):
    today = date.today()
    filename="TradingData.xlsx"
    if not os.path.exists("sub"):
        os.makedirs("sub")
    if os.path.exists(f"{os.path.abspath(__file__)}/sub/{filename}"):
        wb=load_workbook(f"{os.path.abspath(__file__)}/sub/{filename}")
    else:
        wb=Workbook( )
    ws=wb.active

    if ws.cell(row=1,column=ws.max_column).value==str(today):
        pos=ws.max_column
        ws.delete_cols(pos)
    else:
        pos=ws.max_column+1
    ws.cell(row=1,column=pos).value=str(today)    
    ws.title="Trading Data"
    ws["A1"]="Coin"
    for key,value in coins.items():
        found=False
        for i in range(1,ws.max_row):
            if ws.cell(row=i,column=1).value==key:
                ws.cell(row=i,column=pos).value=value
                found=True
        if not found:
            pr=ws.max_row+1
            ws.cell(row=pr,column=1).value=key
            ws.cell(row=pr,column=pos).value=value

    wb.save(f"sub/{filename}") 
